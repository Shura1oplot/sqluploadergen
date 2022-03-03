#!/usr/bin/env python3

import sys
import os
import tempfile
import shutil
import subprocess
import openpyxl


MAP_TYPES = {
    "bigint": "Int64",
    "bit": "bool",
    "char": "string",
    "date": "DateTime",
    "datetime": "DateTime",
    "datetime2": "DateTime",
    "datetimeoffset": "DateTimeOffset",
    "decimal": "decimal",
    "float": "double",
    "int": "Int32",
    "money": "decimal",
    "nchar": "string",
    "ntext": "string",
    "numeric": "decimal",
    "nvarchar": "string",
    "real": "single",
    "smalldatetime": "DateTime",
    "smallint": "Int16",
    "smallmoney": "decimal",
    "text": "string",
    "time": "TimeSpan",
    "tinyint": "byte",
    "varchar": "string",
}

MAP_CONVS = {
    "byte": "ConvertStringToByte",
    "Int16": "ConvertStringToShort",
    "Int32": "ConvertStringToInt",
    "Int64": "ConvertStringToLong",
    "single": "ConvertStringToFloat",
    "double": "ConvertStringToDouble",
    "decimal": "ConvertStringToDecimal",
    "string":  "ConvertStringToString",
    "DateTime": "ConvertStringToDateTime",
    "TimeSpan": "ConvertStringToTimeSpan",

    # not implemented:
    # "DateTimeOffset": "",
}


def main(argv=sys.argv):
    if getattr(sys, "frozen", False):
        prog_name = os.path.basename(sys.executable)
        prog_dir = os.path.dirname(sys.executable)
    else:
        prog_name = os.path.basename(argv[0])
        prog_dir = os.path.dirname(argv[0])

    if len(argv) != 2:
        print(f"usage: {prog_name} specfile.xlsx", file=sys.stderr)
        return 1

    spec_file_name = argv[1]

    spec_file_dir = os.path.dirname(spec_file_name)

    wb = openpyxl.load_workbook(spec_file_name, read_only=True, data_only=True)

    spec = wb["_spec_"]

    server = spec["B2"].value
    database = spec["B3"].value
    domain_auth = spec["B4"].value
    login = spec["B5"].value
    password = spec["B6"].value
    delimiter = spec["B7"].value

    if not server:
        print("error: 'Server' must be specified", file=sys.stderr)
        return 1

    if not database:
        print("error: 'Database' must be specified", file=sys.stderr)
        return 1

    if not isinstance(domain_auth, bool):
        print("error: 'Domain auth' must be TRUE or FALSE", file=sys.stderr)
        return 1

    if not domain_auth and not login:
        print("error: 'Login' must be specified for non-domain auth", file=sys.stderr)
        return 1

    if not domain_auth and not password:
        print("error: 'Password' must be specified for non-domain auth", file=sys.stderr)
        return 1

    if not isinstance(delimiter, str) and len(delimiter) != 1:
        print("error: 'Delimiter' must be one char", file=sys.stderr)
        return 1

    print(f"Server: {server}")
    print(f"Database: {database}")
    print(f"Domain auth: {domain_auth}")
    print(f"Login: {login}")
    print(f"Password: {password}")
    print(f"Delimiter: {delimiter}")

    conn_str = f"Server={server};Database={database};"

    if domain_auth:
        conn_str += "Trusted_Connection=True"
    else:
        conn_str += f"User ID={login};Password={password};"

    tables = {}

    for sheet_name in wb.sheetnames:
        if sheet_name == "_spec_":
            continue

        ws = wb[sheet_name]

        table = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            col_name = row[0]
            col_type = row[1]

            try:
                col_type_param1 = row[2]
            except IndexError:
                col_type_param1 = None

            try:
                col_type_param2 = row[3]
            except IndexError:
                col_type_param2 = None

            table.append((col_name, col_type, col_type_param1, col_type_param2))

        tables[sheet_name] = table

    # FIXME: check table definition

    sql_file = os.path.join(spec_file_dir, "1_init_db.sql")

    with open(sql_file, "w") as fp:
        fp.write(f"USE [{database}]\nGO\n\n")

        for table_name, table_def in tables.items():
            fp.write(f"DROP TABLE IF EXISTS [{table_name}]\nGO\n\n")

            fp.write(f"CREATE TABLE [{table_name}] (\n")

            for i, (col_name, col_type, col_type_param1, col_type_param2) in enumerate(table_def):
                if i == 0:
                    fp.write("    ")
                else:
                    fp.write("   ,")

                fp.write(f"[{col_name}] [{col_type}]")

                if col_type_param1 is not None:
                    fp.write(f"({col_type_param1}")

                    if col_type_param2 is not None:
                        fp.write(f", {col_type_param2}")

                    fp.write(")")

                fp.write("\n")

            fp.write(")\nGO\n\n")

            fp.write(f"-- CREATE CLUSTERED COLUMNSTORE INDEX ix_columnstore "
                     f"ON {table_name}\nGO\n\n")

    tmp_dir = tempfile.mkdtemp()

    for table_name, table_def in tables.items():
        shutil.copytree(
            os.path.join(prog_dir, "assets", "bcpstream_template"),
            os.path.join(tmp_dir, f"bcpstream_{table_name}"))

        Program_cs_file = os.path.join(tmp_dir, f"bcpstream_{table_name}", "Program.cs")
        Program_cs_content = open(Program_cs_file, "r", encoding="utf-8").read()
        Program_cs_content = Program_cs_content.replace("{{connection_string}}", conn_str)
        open(Program_cs_file, "w", encoding="utf-8").write(Program_cs_content)

        cs_columns = []

        for i, (col_name, col_type, _, _) in enumerate(table_def):
            cs_type = MAP_TYPES[col_type]
            nullable = "?" if cs_type != "string" else ""

            cs_columns.append(
                f'        [Column("{col_name}")] public {cs_type}{nullable} '
                f'Column{i} {{ get; set; }}\n')

        cs_columns = "".join(cs_columns)

        cs_converters = []

        for i, (col_name, col_type, _, _) in enumerate(table_def):
            cs_type = MAP_TYPES[col_type]
            conv = MAP_CONVS[cs_type]

            cs_converters.append(
                f'            Column{i} = Helpers.{conv}(values[{i}]);\n')

        cs_converters = "".join(cs_converters)

        Format_cs_file = os.path.join(tmp_dir, f"bcpstream_{table_name}", "Format.cs")
        Format_cs_content = open(Format_cs_file, "r", encoding="utf-8").read()
        Format_cs_content = Format_cs_content.replace("{{table}}", table_name)
        Format_cs_content = Format_cs_content.replace("{{columns}}", cs_columns)
        Format_cs_content = Format_cs_content.replace("{{converters}}", cs_converters)
        Format_cs_content = Format_cs_content.replace("{{delimiter}}", delimiter)
        open(Format_cs_file, "w", encoding="utf-8").write(Format_cs_content)

        subprocess.run(
            args=["dotnet", "publish", "--nologo", "--configuration", "Release",
                  "--self-contained", "false", "--runtime", "win-x64",
                  "/p:PublishSingleFile=true",
                  "/p:IncludeNativeLibrariesForSelfExtract=true"],
            cwd=os.path.join(tmp_dir, f"bcpstream_{table_name}"))

        shutil.copy(
            os.path.join(tmp_dir, f"bcpstream_{table_name}", "bin", "Release",
                         "net5.0", "win-x64", "publish", "bcpstream.exe"),
            os.path.join(spec_file_dir, f"bcpstream_{table_name}.exe"))

    shutil.rmtree(tmp_dir)

    shutil.copy(os.path.join(prog_dir, "assets", "tools", "etl.exe"),
                os.path.join(spec_file_dir, "etl.exe"))
    shutil.copy(os.path.join(prog_dir, "assets", "tools", "7za.exe"),
                os.path.join(spec_file_dir, "7za.exe"))

    with open(os.path.join(spec_file_dir, "2_upload.bat"),
              "w", encoding="cp866") as fp:
        fp.write("@ECHO OFF\n\n")

        for i, table_name in enumerate(tables, start=1):
            fp.write(f'7za e data.zip -so data.csv | etl decode "UTF-8" '
                     f'| etl selectrows 2 '
                     f'| bcpstream_{table_name} "{i}/{len(tables)}" 111111111\n')


if __name__ == "__main__":
    sys.exit(main())
